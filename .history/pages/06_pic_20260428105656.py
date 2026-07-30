import streamlit as st
import pandas as pd
import datetime as dt
import re
from io import BytesIO
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from shared import get_engine, logo_sidebar, wk_cols_from_df

st.title("📅 PIC — Agrégation Mensuelle")

if get_engine() is None:
    st.stop()

with st.sidebar:
    logo_sidebar()
    st.header("⚙️ Info")
    st.markdown("---")
    st.info(
        "Même vue que la Consolidée mais agrégée par **mois**.\n\n"
        "Vous pouvez aussi uploader un fichier PIC exporté "
        "et modifié manuellement.\n\n"
        "👈 Lancez d'abord la page **📊 Consolidée** pour alimenter cette vue."
    )

# ── Source ────────────────────────────────────────────────────────────────────
source = st.radio(
    "📁 Source",
    ["🔗 Session (Agrégation)", "📂 Fichier Excel modifié"],
    horizontal=True
)

def wk_to_month(label):
    try:
        yy = int('20' + label[1:3])
        ww = int(label[4:6])
        return dt.date.fromisocalendar(yy, ww, 1).strftime('%Y-%m')
    except:
        return None

def month_label(ym):
    try:
        return dt.datetime.strptime(ym, '%Y-%m').strftime('%b-%y')
    except:
        return ym

COL_GROUPE = 'SERTA_SO_CLIENT_GROUP_NAME'
COL_PRIX   = 'PRIX_MOQ'

META_COLS = ['REF_ARTICLE_SERTA', 'REF_ARTICLE_CLIENT', 'ORIGINE', 'UP_PRINCIPALE',
             'CODE_SELECTION', 'QTE_MOQ', 'QTE_UC', 'PROGRAMME', 'HORIZON_PROGRAMME',
             'PRIX_MOQ', 'SERTA_SO_CLIENT_GROUP_NAME', 'SERTA_SO_CLIENT_NAME',
             'SALES_ADMINISTRATION_PERSON', 'CODE_CLIENT', 'DOUBLON']

if source.startswith("🔗"):
    if 'df_consolide' not in st.session_state:
        st.warning("⚠️ Lancez d'abord la page **📊 Consolidée** pour alimenter cette vue.")
        st.stop()

    df_raw = st.session_state['df_consolide'].copy()
    wk_cols = wk_cols_from_df(df_raw)

    # Forcer numériques semaines
    for c in wk_cols:
        df_raw[c] = pd.to_numeric(df_raw[c], errors='coerce').fillna(0)

    # Construire mois map
    mois_map = {}
    for wk in wk_cols:
        m = wk_to_month(wk)
        if m:
            mois_map.setdefault(m, []).append(wk)
    mois_tries  = sorted(mois_map.keys())
    mois_labels = [month_label(m) for m in mois_tries]

    # Agréger semaines → mois par ligne (garder toutes les métadonnées)
    meta_pres = [c for c in META_COLS if c in df_raw.columns]
    df_pic = df_raw[meta_pres].copy()
    for m, wks in mois_map.items():
        ml = month_label(m)
        df_pic[ml] = df_raw[[w for w in wks if w in df_raw.columns]].sum(axis=1)

    # ── Appliquer cutoff sur le mois de date_prevision ──────────────────────────
    _date_prev = st.session_state.get('date_prevision', None)
    if _date_prev:
        _mois_prev = _date_prev.strftime('%Y-%m')
        _ml_prev   = dt.datetime.strptime(_mois_prev, '%Y-%m').strftime('%b-%y')
        for _ml in mois_labels:
            try:
                _d = dt.datetime.strptime(_ml, '%b-%y')
                if _d.strftime('%Y-%m') < _mois_prev:
                    # Mois passés → 0
                    df_pic[_ml] = 0
                elif _ml == _ml_prev:
                    # Mois cutoff → besoin encours uniquement
                    _lpc_encours = pd.to_numeric(df_raw.get('QTE_BESOIN_CLIENT_ENCOURS',    0), errors='coerce').fillna(0)
                    _sc_encours  = pd.to_numeric(df_raw.get('QTE_BESOIN_CLIENT_ENCOURS_SC', 0), errors='coerce').fillna(0)
                    df_pic[_ml]  = pd.Series((_lpc_encours + _sc_encours).values).clip(lower=0).values
            except:
                pass

    df_pic['TOTAL QTY'] = df_pic[mois_labels].sum(axis=1)
    st.success(f"✅ {len(df_pic)} lignes depuis la session")

else:
    uploaded = st.file_uploader(
        "Uploader le fichier PIC Excel modifié",
        type=['xlsx', 'xls'],
        help="Même format que l'export PIC — colonnes méta + colonnes mois (Apr-26, May-26...)"
    )
    if not uploaded:
        st.info("👆 Uploadez le fichier PIC Excel exporté et modifié.")
        st.stop()
    try:
        df_pic = pd.read_excel(uploaded)
        mois_labels = [c for c in df_pic.columns if re.match(r'^[A-Za-z]{3}-\d{2}$', str(c))]
        for c in mois_labels:
            df_pic[c] = pd.to_numeric(df_pic[c], errors='coerce').fillna(0)
        if 'TOTAL QTY' not in df_pic.columns:
            df_pic['TOTAL QTY'] = df_pic[mois_labels].sum(axis=1)
        else:
            df_pic['TOTAL QTY'] = pd.to_numeric(df_pic['TOTAL QTY'], errors='coerce').fillna(0)
        meta_pres = [c for c in META_COLS if c in df_pic.columns]
        st.success(f"✅ {len(df_pic)} lignes chargées — {len(mois_labels)} mois")
    except Exception as e:
        st.error(f"❌ {e}")
        st.stop()

st.markdown("---")

# ── Filtre groupe client ──────────────────────────────────────────────────────
if COL_GROUPE in df_pic.columns:
    groupes_dispo = sorted([g for g in df_pic[COL_GROUPE].dropna().astype(str).unique() if g.strip()])
    f_groupes = st.multiselect("🏢 Filtrer par groupe client", options=groupes_dispo,
                                placeholder="— tous les groupes —")
    if f_groupes:
        df_pic = df_pic[
            df_pic[COL_GROUPE].astype(str).str.strip().isin(f_groupes)
        ]

# ── Filtres ───────────────────────────────────────────────────────────────────
with st.expander("🔍 Filtres", expanded=False):
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        f_ref = st.multiselect("Ref SERTA",
            options=sorted(df_pic['REF_ARTICLE_SERTA'].dropna().astype(str).unique())
            if 'REF_ARTICLE_SERTA' in df_pic.columns else [])
    with col2:
        f_up = st.multiselect("UP",
            options=sorted(df_pic['UP_PRINCIPALE'].dropna().replace('', None).dropna().astype(str).unique())
            if 'UP_PRINCIPALE' in df_pic.columns else [])
    with col3:
        f_orig = st.multiselect("Origine",
            options=sorted(df_pic['ORIGINE'].dropna().replace('', None).dropna().astype(str).unique())
            if 'ORIGINE' in df_pic.columns else [])
    with col4:
        f_prog = st.multiselect("Programme",
            options=sorted(df_pic['PROGRAMME'].dropna().replace('', None).dropna().astype(str).unique())
            if 'PROGRAMME' in df_pic.columns else [])

df_disp = df_pic.copy()
if f_ref:  df_disp = df_disp[df_disp['REF_ARTICLE_SERTA'].astype(str).isin(f_ref)]
if f_up:   df_disp = df_disp[df_disp['UP_PRINCIPALE'].astype(str).isin(f_up)]
if f_orig: df_disp = df_disp[df_disp['ORIGINE'].astype(str).isin(f_orig)]
if f_prog: df_disp = df_disp[df_disp['PROGRAMME'].astype(str).isin(f_prog)]

# ── Métriques ─────────────────────────────────────────────────────────────────
c1, c2, c3, c4 = st.columns(4)
c1.metric("Refs",        len(df_disp))
c2.metric("Mois",        len(mois_labels))
c3.metric("QTY Total",   f"{int(df_disp['TOTAL QTY'].sum()):,}")
if COL_GROUPE in df_disp.columns:
    c4.metric("Groupes clients", df_disp[COL_GROUPE].nunique())
if mois_labels:
    st.caption(f"📅 {mois_labels[0]} → {mois_labels[-1]}")

# ── Tableau ───────────────────────────────────────────────────────────────────
col_cfg = {m: st.column_config.NumberColumn(m, format="%d") for m in mois_labels + ['TOTAL QTY']}
cols_affich = [c for c in meta_pres if c in df_disp.columns] + mois_labels + ['TOTAL QTY']
st.caption(f"{len(df_disp):,} lignes")
st.dataframe(df_disp[[c for c in cols_affich if c in df_disp.columns]],
             width='stretch', height=600, column_config=col_cfg)

st.markdown("---")

# ── Export Excel ──────────────────────────────────────────────────────────────
from datetime import datetime as _dt2
buf = BytesIO()
with pd.ExcelWriter(buf, engine='openpyxl') as writer:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def style_ws(ws, color='1F4E79'):
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
            cell.fill = PatternFill('solid', start_color=color)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 30
        fill_alt = PatternFill('solid', start_color='EBF3FF')
        thin = Border(left=Side(style='thin', color='BFBFBF'),
                      right=Side(style='thin', color='BFBFBF'),
                      bottom=Side(style='thin', color='BFBFBF'))
        for i, row in enumerate(ws.iter_rows(min_row=2), 2):
            for cell in row:
                cell.font = Font(name='Arial', size=10)
                cell.border = thin
                if i % 2 == 0:
                    cell.fill = fill_alt
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')
        for col in ws.columns:
            w = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, 20)
        ws.freeze_panes = ws.cell(row=2, column=len(meta_pres) + 1)

    export_cols = [c for c in cols_affich if c in df_disp.columns]
    df_disp[export_cols].to_excel(writer, sheet_name='PIC Mensuel', index=False)
    style_ws(writer.sheets['PIC Mensuel'], '1F4E79')

st.download_button(
    "📥 Télécharger le PIC Excel",
    data=buf.getvalue(),
    file_name=f"PIC_{_dt2.now().strftime('%Y%m%d_%H%M')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True
)

# ── Sauvegarde SQL ────────────────────────────────────────────────────────────
st.markdown("---")
st.subheader("🗄️ Archiver en base SQL")
col_sv1, col_sv2 = st.columns([3, 1])
with col_sv1:
    commentaire = st.text_input("Commentaire", placeholder="ex: PIC Avril 2026 validé")
with col_sv2:
    btn_save = st.button("💾 Sauvegarder", type="primary", use_container_width=True)

if btn_save:
    from sqlalchemy import text as _text
    engine = get_engine()
    if engine is None:
        st.error("❌ Connexion SQL indisponible")
    else:
        try:
            date_export = _dt2.now().date()
            id_cols = [c for c in meta_pres if c in df_disp.columns]
            df_long = df_disp[id_cols + mois_labels].melt(
                id_vars=id_cols, var_name='MOIS_LABEL', value_name='QTY'
            )
            df_long['ANNEE_MOIS'] = df_long['MOIS_LABEL'].apply(
                lambda x: dt.datetime.strptime(x, '%b-%y').strftime('%Y-%m')
                if re.match(r'^[A-Za-z]{3}-\d{2}$', str(x)) else None
            )
            insert_q = _text("""INSERT INTO [master].[dbo].[T_PIC_MENSUEL]
                (DATE_EXPORT,ANNEE_MOIS,MOIS_LABEL,GROUPE_CLIENT,NOM_CLIENT,
                 REF_ARTICLE_SERTA,UP_PRINCIPALE,ORIGINE,PROGRAMME,QTY,COMMENTAIRE)
                VALUES(:de,:am,:ml,:gc,:nc,:ref,:up,:orig,:prog,:qty,:com)""")
            with engine.begin() as conn:
                for _, row in df_long.iterrows():
                    conn.execute(insert_q, {
                        'de':   date_export,
                        'am':   row.get('ANNEE_MOIS'),
                        'ml':   row.get('MOIS_LABEL'),
                        'gc':   row.get(COL_GROUPE),
                        'nc':   row.get('SERTA_SO_CLIENT_NAME'),
                        'ref':  row.get('REF_ARTICLE_SERTA'),
                        'up':   row.get('UP_PRINCIPALE'),
                        'orig': row.get('ORIGINE'),
                        'prog': row.get('PROGRAMME'),
                        'qty':  row.get('QTY'),
                        'com':  commentaire or None,
                    })
            st.success(f"✅ {len(df_long):,} lignes archivées (export du {date_export})")
        except Exception as e:
            st.error(f"❌ {e}")
            import traceback
            with st.expander("Détails"):
                st.code(traceback.format_exc())