import streamlit as st
import pandas as pd
import datetime as dt
import re
from io import BytesIO
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from shared import get_engine, logo_sidebar, wk_cols_from_df

st.title("📅 PIC — Agrégation Mensuelle")

if get_engine() is None:
    st.stop()

with st.sidebar:
    logo_sidebar()
    st.header("⚙️ Info")
    st.markdown("---")
    st.info(
        "Même vue que la Consolidée mais agrégée par **mois**.\n\n"
        "Vous pouvez aussi uploader un fichier PIC exporté "
        "et modifié manuellement.\n\n"
        "👈 Lancez d'abord la page **📊 Consolidée** pour alimenter cette vue."
    )
    st.markdown("---")
    # AJOUT : filtre PDR/composants, absent jusqu'ici du PIC (existait deja
    # en page 04 Consolidee). Meme principe : desactive par defaut, exclut
    # les refs classees PDR/PFPDR/MAUNIT/MAJOIN (vrai CODE_GROUPE_ARTICLE).
    filtrer_pdr_pic = st.checkbox(
        "Filtrer les PDR/composants",
        value=False,
        help="Désactivé par défaut. Exclut les références classées "
             "PDR/PFPDR/MAUNIT/MAJOIN (vrai CODE_GROUPE_ARTICLE) -- "
             "même filtre que celui de la page Consolidée.")

# ── Source ────────────────────────────────────────────────────────────────────
source = st.radio(
    "📁 Source",
    ["🔗 Session (Agrégation)", "📂 Fichier Excel modifié"],
    horizontal=True
)

def wk_to_month(label):
    try:
        yy = int('20' + label[1:3])
        ww = int(label[4:6])
        return dt.date.fromisocalendar(yy, ww, 1).strftime('%Y-%m')
    except:
        return None

# AJOUT : mapping explicite anglais->numero de mois, independant de la locale
# systeme. dt.datetime.strptime(..., '%b-%y') depend de la langue configuree
# sur le serveur -- si le systeme est en francais (janv./fevr. au lieu de
# Jan/Feb), le parsing echoue SILENCIEUSEMENT (cache par le except: pass plus
# bas), et tout le calcul de cutoff/trimestre ne se declenche jamais sans
# aucun message d'erreur visible.
MOIS_ABBREV_EN = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,
                   'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}

def parse_month_label(ml):
    """Parse 'Jan-27' -> date(2027,1,1), sans dependance a la locale systeme."""
    try:
        abbrev, yy = ml.split('-')
        mm = MOIS_ABBREV_EN[abbrev]
        yyyy = 2000 + int(yy)
        return dt.date(yyyy, mm, 1)
    except (KeyError, ValueError):
        return None

def month_label(ym):
    try:
        d = dt.datetime.strptime(ym, '%Y-%m')
        abbrev = [k for k,v in MOIS_ABBREV_EN.items() if v == d.month][0]
        return f"{abbrev}-{str(d.year)[2:]}"
    except:
        return ym

def month_to_quarter_label(ym):
    """'2026-07' -> 'Q3-26'"""
    try:
        d = dt.datetime.strptime(ym, '%Y-%m')
        q = (d.month - 1) // 3 + 1
        return f"Q{q}-{str(d.year)[2:]}"
    except:
        return None

COL_GROUPE = 'SERTA_SO_CLIENT_GROUP_NAME'
COL_PRIX   = 'PRIX_MOQ'

META_COLS = ['REF_ARTICLE_SERTA', 'REF_ARTICLE_CLIENT', 'ORIGINE', 'UP_PRINCIPALE',
             'CODE_SELECTION', 'QTE_MOQ', 'QTE_UC', 'PROGRAMME', 'HORIZON_PROGRAMME',
             'PRIX_MOQ', 'SERTA_SO_CLIENT_GROUP_NAME', 'SERTA_SO_CLIENT_NAME',
             'SALES_ADMINISTRATION_PERSON', 'CODE_CLIENT', 'DOUBLON']

if source.startswith("🔗"):
    # CORRIGE : cette page lisait 'df_consolide', la cle que la page 04
    # ecrivait AVANT le correctif de collision de session -- depuis ce
    # correctif, la page 04 ecrit dans 'df_consolide_final' (pour ne plus
    # ecraser le programme brut de la page 02, qui utilise aussi la cle
    # 'df_consolide'). Sans cette mise a jour, le PIC lisait le programme
    # brut fige, jamais enrichi par le carnet ni les projets -- d'ou
    # l'absence totale de refs CARNET/PROJET observee dans les exports PIC.
    if 'df_consolide_final' not in st.session_state:
        st.warning("⚠️ Lancez d'abord la page **📊 Consolidée** pour alimenter cette vue.")
        st.stop()

    df_raw = st.session_state['df_consolide_final'].copy()
    wk_cols = wk_cols_from_df(df_raw)

    # Forcer numériques semaines
    for c in wk_cols:
        df_raw[c] = pd.to_numeric(df_raw[c], errors='coerce').fillna(0)

    # Construire mois map
    mois_map = {}
    for wk in wk_cols:
        m = wk_to_month(wk)
        if m:
            mois_map.setdefault(m, []).append(wk)
    mois_tries  = sorted(mois_map.keys())
    mois_labels = [month_label(m) for m in mois_tries]

    # Agréger semaines → mois par ligne (garder toutes les métadonnées)
    meta_pres = [c for c in META_COLS if c in df_raw.columns]
    df_pic = df_raw[meta_pres].copy()
    for m, wks in mois_map.items():
        ml = month_label(m)
        df_pic[ml] = df_raw[[w for w in wks if w in df_raw.columns]].sum(axis=1)

    # ── Appliquer cutoff sur le mois de date_prevision ──────────────────────────
    _date_prev = st.session_state.get('date_prevision', None)
    if _date_prev:
        _mois_prev = _date_prev.strftime('%Y-%m')
        _ml_prev   = month_label(_mois_prev)
        _erreurs_parsing = []
        for _ml in mois_labels:
            _d = parse_month_label(_ml)
            if _d is None:
                _erreurs_parsing.append(_ml)
                continue
            _ym = _d.strftime('%Y-%m')
            if _ym < _mois_prev:
                # Mois passés → 0
                df_pic[_ml] = 0
            elif _ml == _ml_prev:
                # CORRIGE : le remplacement par l'encours ne doit s'appliquer
                # qu'aux lignes qui touchent au CARNET -- avant, TOUTES les
                # lignes (y compris LPC pur, sans commande ferme) etaient
                # remplacees par l'encours, qui vaut 0 pour elles puisque ces
                # colonnes viennent specifiquement du carnet. Consequence :
                # toute prevision LPC pure sur le mois de cutoff se
                # retrouvait mise a 0 a tort, alors qu'elle represente une
                # vraie prevision programme, pas juste "pas de commande ferme".
                _lpc_encours = pd.to_numeric(df_raw.get('QTE_BESOIN_CLIENT_ENCOURS',    0), errors='coerce').fillna(0)
                _sc_encours  = pd.to_numeric(df_raw.get('QTE_BESOIN_CLIENT_ENCOURS_SC', 0), errors='coerce').fillna(0)
                _encours_totale = pd.Series((_lpc_encours + _sc_encours).values).clip(lower=0).values

                if 'ORIGINE' in df_raw.columns:
                    _touche_carnet = df_raw['ORIGINE'].astype(str).str.contains('CARNET', na=False).values
                    df_pic.loc[_touche_carnet, _ml] = _encours_totale[_touche_carnet]
                    # Lignes LPC pures (sans CARNET) : la valeur du mois deja
                    # calculee (prevision programme) reste inchangee.
                else:
                    # ORIGINE indisponible -- comportement precedent par defaut
                    df_pic[_ml] = _encours_totale
        # AJOUT : avertissement visible au lieu d'un echec silencieux
        # (l'ancien code avait un except: pass qui masquait toute erreur de
        # parsing, notamment le risque de locale systeme aborde plus haut)
        if _erreurs_parsing:
            st.warning(f"⚠️ {len(_erreurs_parsing)} libellé(s) de mois n'ont pas pu être "
                       f"interprétés pour le cutoff : {_erreurs_parsing}")
    else:
        st.caption("ℹ️ Pas de cutoff mensuel appliqué (date_prevision non définie en page 01).")

    df_pic['TOTAL QTY'] = df_pic[mois_labels].sum(axis=1)
    st.success(f"✅ {len(df_pic)} lignes depuis la session")

else:
    uploaded = st.file_uploader(
        "Uploader le fichier PIC Excel modifié",
        type=['xlsx', 'xls'],
        help="Même format que l'export PIC — colonnes méta + colonnes mois (Apr-26, May-26...)"
    )
    if not uploaded:
        st.info("👆 Uploadez le fichier PIC Excel exporté et modifié.")
        st.stop()
    try:
        df_pic = pd.read_excel(uploaded)
        mois_labels = [c for c in df_pic.columns if re.match(r'^[A-Za-z]{3}-\d{2}$', str(c))]
        for c in mois_labels:
            df_pic[c] = pd.to_numeric(df_pic[c], errors='coerce').fillna(0)
        if 'TOTAL QTY' not in df_pic.columns:
            df_pic['TOTAL QTY'] = df_pic[mois_labels].sum(axis=1)
        else:
            df_pic['TOTAL QTY'] = pd.to_numeric(df_pic['TOTAL QTY'], errors='coerce').fillna(0)
        meta_pres = [c for c in META_COLS if c in df_pic.columns]
        st.success(f"✅ {len(df_pic)} lignes chargées — {len(mois_labels)} mois")
    except Exception as e:
        st.error(f"❌ {e}")
        st.stop()

st.markdown("---")

# AJOUT : filtre PDR/composants, applique ici -- avant le calcul des
# trimestres pour que ceux-ci refletent aussi le filtre. Meme fonction/
# logique que celle de la page Consolidee (echappement de quotes corrige,
# lots de 100 refs, sans cache st.cache_data pour eviter les soucis
# d'affichage UI dans une fonction cachee).
if filtrer_pdr_pic and 'REF_ARTICLE_SERTA' in df_pic.columns:
    from sqlalchemy import text as _text_pdr

    def _charger_groupe_article_pic(refs_tuple):
        engine = get_engine()
        if engine is None or not refs_tuple:
            return pd.DataFrame(columns=['REF_ARTICLE_SERTA', 'VRAI_GROUPE_ARTICLE'])
        TAILLE_LOT = 100
        lots = [refs_tuple[i:i+TAILLE_LOT] for i in range(0, len(refs_tuple), TAILLE_LOT)]
        resultats = []
        progress = st.progress(0, text="Vérification du groupe article...")
        for i, lot in enumerate(lots):
            refs_echappees = ["''" + r.replace("'", "''''") + "''" for r in lot]
            refs_sql = ", ".join(refs_echappees)
            try:
                with engine.connect() as conn:
                    df_lot = pd.read_sql(_text_pdr(f"""
                        SELECT *
                        FROM OPENQUERY([SRV-MSSQLDB], '
                            SELECT REF_ARTICLE, CODE_GROUPE_ARTICLE
                            FROM DW.PRODUCTION.V_ART_STANDARD
                            WHERE REF_ARTICLE IN ({refs_sql})
                        ')
                    """), conn)
                df_lot.columns = ['REF_ARTICLE_SERTA', 'VRAI_GROUPE_ARTICLE']
                resultats.append(df_lot)
            except Exception:
                pass
            progress.progress((i+1)/len(lots), text=f"Vérification groupe article... lot {i+1}/{len(lots)}")
        progress.empty()
        if not resultats:
            return pd.DataFrame(columns=['REF_ARTICLE_SERTA', 'VRAI_GROUPE_ARTICLE'])
        df_res = pd.concat(resultats, ignore_index=True)
        df_res['REF_ARTICLE_SERTA'] = df_res['REF_ARTICLE_SERTA'].astype(str).str.strip()
        return df_res.drop_duplicates('REF_ARTICLE_SERTA')

    GROUPES_PDR_PIC = ['PDR', 'PFPDR', 'MAUNIT', 'MAJOIN']
    refs_pic_tuple = tuple(sorted(df_pic['REF_ARTICLE_SERTA'].dropna().astype(str).str.strip().unique()))
    df_groupe_pic = _charger_groupe_article_pic(refs_pic_tuple)
    if not df_groupe_pic.empty:
        nb_avant_pic = len(df_pic)
        df_pic = df_pic.merge(df_groupe_pic, on='REF_ARTICLE_SERTA', how='left')
        a_retirer_pic = df_pic['VRAI_GROUPE_ARTICLE'].isin(GROUPES_PDR_PIC)
        nb_retirees_pic = a_retirer_pic.sum()
        df_pic = df_pic[~a_retirer_pic].drop(columns=['VRAI_GROUPE_ARTICLE'])
        if nb_retirees_pic > 0:
            st.info(f"🧹 {nb_retirees_pic} ligne(s) retirée(s) (PDR/composants).")

# AJOUT : agregation trimestrielle -- absente du fichier original, construite
# a partir des colonnes mois deja calculees (meme logique, lundi de la
# semaine ISO determine le mois, donc le trimestre par transitivite).
trimestre_map = {}
for ml in mois_labels:
    d = parse_month_label(ml)
    if d is None:
        continue
    ql = month_to_quarter_label(d.strftime('%Y-%m'))
    if ql:
        trimestre_map.setdefault(ql, []).append(ml)

def _tri_sort_key(ql):
    q, yy = ql.split('-')
    return (int(yy), int(q[1]))
trimestres_tries = sorted(trimestre_map.keys(), key=_tri_sort_key)

for ql in trimestres_tries:
    mois_du_trimestre = [m for m in trimestre_map[ql] if m in df_pic.columns]
    df_pic[ql] = df_pic[mois_du_trimestre].sum(axis=1)

if trimestres_tries:
    st.caption(f"📅 Trimestres disponibles : {trimestres_tries[0]} → {trimestres_tries[-1]} "
               f"({len(trimestres_tries)} trimestres)")

# ── Filtre groupe client ──────────────────────────────────────────────────────
if COL_GROUPE in df_pic.columns:
    groupes_dispo = sorted([g for g in df_pic[COL_GROUPE].dropna().astype(str).unique() if g.strip()])
    f_groupes = st.multiselect("🏢 Filtrer par groupe client", options=groupes_dispo,
                                placeholder="— tous les groupes —")
    if f_groupes:
        df_pic = df_pic[
            df_pic[COL_GROUPE].astype(str).str.strip().isin(f_groupes)
        ]

# ── Filtres ───────────────────────────────────────────────────────────────────
with st.expander("🔍 Filtres", expanded=False):
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        f_ref = st.multiselect("Ref SERTA",
            options=sorted(df_pic['REF_ARTICLE_SERTA'].dropna().astype(str).unique())
            if 'REF_ARTICLE_SERTA' in df_pic.columns else [])
    with col2:
        f_up = st.multiselect("UP",
            options=sorted(df_pic['UP_PRINCIPALE'].dropna().replace('', None).dropna().astype(str).unique())
            if 'UP_PRINCIPALE' in df_pic.columns else [])
    with col3:
        f_orig = st.multiselect("Origine",
            options=sorted(df_pic['ORIGINE'].dropna().replace('', None).dropna().astype(str).unique())
            if 'ORIGINE' in df_pic.columns else [])
    with col4:
        f_prog = st.multiselect("Programme",
            options=sorted(df_pic['PROGRAMME'].dropna().replace('', None).dropna().astype(str).unique())
            if 'PROGRAMME' in df_pic.columns else [])

df_disp = df_pic.copy()
if f_ref:  df_disp = df_disp[df_disp['REF_ARTICLE_SERTA'].astype(str).isin(f_ref)]
if f_up:   df_disp = df_disp[df_disp['UP_PRINCIPALE'].astype(str).isin(f_up)]
if f_orig: df_disp = df_disp[df_disp['ORIGINE'].astype(str).isin(f_orig)]
if f_prog: df_disp = df_disp[df_disp['PROGRAMME'].astype(str).isin(f_prog)]

# ── Métriques ─────────────────────────────────────────────────────────────────
c1, c2, c3, c4 = st.columns(4)
c1.metric("Refs",        len(df_disp))
c2.metric("Mois",        len(mois_labels))
c3.metric("QTY Total",   f"{int(df_disp['TOTAL QTY'].sum()):,}")
if COL_GROUPE in df_disp.columns:
    c4.metric("Groupes clients", df_disp[COL_GROUPE].nunique())
if mois_labels:
    st.caption(f"📅 {mois_labels[0]} → {mois_labels[-1]}")

# ── Tableau ───────────────────────────────────────────────────────────────────
col_cfg = {m: st.column_config.NumberColumn(m, format="%d") for m in mois_labels + trimestres_tries + ['TOTAL QTY']}
cols_affich = [c for c in meta_pres if c in df_disp.columns] + mois_labels + trimestres_tries + ['TOTAL QTY']
st.caption(f"{len(df_disp):,} lignes")
st.dataframe(df_disp[[c for c in cols_affich if c in df_disp.columns]],
             width='stretch', height=600, column_config=col_cfg)

st.markdown("---")

# ── Export Excel ──────────────────────────────────────────────────────────────
from datetime import datetime as _dt2
buf = BytesIO()
with pd.ExcelWriter(buf, engine='openpyxl') as writer:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def style_ws(ws, color='1F4E79'):
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
            cell.fill = PatternFill('solid', start_color=color)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 30
        fill_alt = PatternFill('solid', start_color='EBF3FF')
        thin = Border(left=Side(style='thin', color='BFBFBF'),
                      right=Side(style='thin', color='BFBFBF'),
                      bottom=Side(style='thin', color='BFBFBF'))
        for i, row in enumerate(ws.iter_rows(min_row=2), 2):
            for cell in row:
                cell.font = Font(name='Arial', size=10)
                cell.border = thin
                if i % 2 == 0:
                    cell.fill = fill_alt
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')
        for col in ws.columns:
            w = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, 20)
        ws.freeze_panes = ws.cell(row=2, column=len(meta_pres) + 1)

    export_cols = [c for c in cols_affich if c in df_disp.columns]
    df_disp[export_cols].to_excel(writer, sheet_name='PIC Mensuel', index=False)
    style_ws(writer.sheets['PIC Mensuel'], '1F4E79')

st.download_button(
    "📥 Télécharger le PIC Excel",
    data=buf.getvalue(),
    file_name=f"PIC_{_dt2.now().strftime('%Y%m%d_%H%M')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    width='stretch'
)

# ── Sauvegarde SQL ────────────────────────────────────────────────────────────
st.markdown("---")
st.subheader("🗄️ Archiver en base SQL")
col_sv1, col_sv2 = st.columns([3, 1])
with col_sv1:
    commentaire = st.text_input("Commentaire", placeholder="ex: PIC Avril 2026 validé")
with col_sv2:
    btn_save = st.button("💾 Sauvegarder", type="primary", width='stretch')

if btn_save:
    from sqlalchemy import text as _text
    engine = get_engine()
    if engine is None:
        st.error("❌ Connexion SQL indisponible")
    else:
        try:
            date_export = _dt2.now().date()
            id_cols = [c for c in meta_pres if c in df_disp.columns]
            df_long = df_disp[id_cols + mois_labels].melt(
                id_vars=id_cols, var_name='MOIS_LABEL', value_name='QTY'
            )
            df_long['ANNEE_MOIS'] = df_long['MOIS_LABEL'].apply(
                lambda x: dt.datetime.strptime(x, '%b-%y').strftime('%Y-%m')
                if re.match(r'^[A-Za-z]{3}-\d{2}$', str(x)) else None
            )
            insert_q = _text("""INSERT INTO [master].[dbo].[T_PIC_MENSUEL]
                (DATE_EXPORT,ANNEE_MOIS,MOIS_LABEL,GROUPE_CLIENT,NOM_CLIENT,
                 REF_ARTICLE_SERTA,UP_PRINCIPALE,ORIGINE,PROGRAMME,QTY,COMMENTAIRE)
                VALUES(:de,:am,:ml,:gc,:nc,:ref,:up,:orig,:prog,:qty,:com)""")
            with engine.begin() as conn:
                for _, row in df_long.iterrows():
                    conn.execute(insert_q, {
                        'de':   date_export,
                        'am':   row.get('ANNEE_MOIS'),
                        'ml':   row.get('MOIS_LABEL'),
                        'gc':   row.get(COL_GROUPE),
                        'nc':   row.get('SERTA_SO_CLIENT_NAME'),
                        'ref':  row.get('REF_ARTICLE_SERTA'),
                        'up':   row.get('UP_PRINCIPALE'),
                        'orig': row.get('ORIGINE'),
                        'prog': row.get('PROGRAMME'),
                        'qty':  row.get('QTY'),
                        'com':  commentaire or None,
                    })
            st.success(f"✅ {len(df_long):,} lignes archivées (export du {date_export})")
        except Exception as e:
            st.error(f"❌ {e}")
            import traceback
            with st.expander("Détails"):
                st.code(traceback.format_exc())